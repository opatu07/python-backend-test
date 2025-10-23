import re # 正規表現機能で使用するためのライブラリ
import sys # システム関連の機能を提供するライブラリ
import pandas as pd # データ操作と分析のためのライブラリ
import pdfplumber # PDFファイルからテキストや表を抽出するためのライブラリ
from pathlib import Path # ファイルパス操作のためのライブラリ
from typing import List, Tuple, Dict, Any, Optional # 型ヒント用のライブラリ
from datetime import datetime, date, timedelta # 日付と時間の操作のためのライブラリ
import calendar # カレンダー関連の機能を提供するライブラリ

INPUT_PDF = "oct.pdf" # 入力PDFファイル名

BUILDING_BOUNDARY_PATTERN = re.compile(r"(１|2|３|１|２|３)(F|Ｆ)$") # 建物の階数を示すパターン
TIME_PAIR_TOKEN_PATTERN = re.compile(r"^○\s*([0-2]?\d:\d{2})\|([0-2]?\d:\d{2})$") # 営業時間のパターン
ONLY_MARK_TOKEN_PATTERN = re.compile(r"^[○×]$") # 営業可否のみを示すパターン
KEYWORDS_NEAR_NAME = ["平日", "月～土", "土", "日～月", "日～土", "月", "火", "水", "木", "金"] # 店舗名の近くに出現する可能性のあるキーワード

def tokens_from_pdf(pdf_path: str) -> List[str]: # PDFからトークンを抽出する関数
    tokens: List[str] = [] # トークンを格納するリスト
    with pdfplumber.open(pdf_path) as pdf: # PDFファイルを開く
        for page in pdf.pages: # 各ページを処理
            words = page.extract_words(use_text_flow=True, keep_blank_chars=False) # ページから単語を抽出
            page_tokens = [w["text"] for w in words] # 単語のテキストをリストに変換
            tokens.extend(page_tokens) # トークンリストに追加

    return tokens # トークンリストを返す

def find_building_hits(tokens: List[str]) -> List[Tuple[int, str]]: # 建物の境界を示すトークンを見つける関数
    hits: List[Tuple[int, str]] = [] # ヒットしたトークンのインデックスとテキストを格納するリスト
    for i, t in enumerate(tokens): # 各トークンを処理
        if t == "本山寮": # 特定の建物名を検出
            hits.append((i, t)); continue # ヒットリストに追加して次へ
        if len(t) >= 2 and BUILDING_BOUNDARY_PATTERN.search(t): # 建物の階数パターンを検出
            hits.append((i, t)) # ヒットリストに追加
    return hits # ヒットリストを返す

def split_blocks_by_hits(tokens: List[str], hits: List[Tuple[int, str]]) -> List[Tuple[str, int, int, List[str]]]: # ヒットしたトークンでブロックを分割する関数
    blocks: List[Tuple[str, int, int, List[str]]] = [] # ブロック情報を格納するリスト
    for i, (idx, tok) in enumerate(hits): # 各ヒットを処理
        start = idx # ブロックの開始インデックス
        end = hits[i + 1][0] if i + 1 < len(hits) else len(tokens) # ブロックの終了インデックス
        block_tokens = tokens[start:end] # ブロック内のトークンを抽出
        blocks.append((tok, start, end, block_tokens)) # ブロック情報をリストに追加
    return blocks # ブロック情報リストを返す

def extract_shop_name_and_base_hours(block_tokens: List[str]) -> Tuple[str, Optional[str]]: # 店舗名と基準営業時間を抽出する関数
    name_tokens: List[str] = [] # 店舗名のトークンを格納するリスト
    kw_idx: Optional[int] = None # キーワードのインデックス
    for i in range(1, min(len(block_tokens), 12)): # 最初の12トークンをチェック
        if any(k in block_tokens[i] for k in KEYWORDS_NEAR_NAME): # キーワードが含まれているかチェック 
            kw_idx = i; break # キーワードのインデックスを設定してループを抜ける

    if kw_idx is None: # キーワードが見つからなかった場合
        name_tokens = block_tokens[1:3] # 最初の2トークンを店舗名とする
        base_hours = None # 基準営業時間はNone
    else: # キーワードが見つかった場合
        name_tokens = block_tokens[1:kw_idx] # キーワードまでのトークンを店舗名とする
        base_hours = None # 基準営業時間の初期値はNone
        for j in range(kw_idx + 1, min(len(block_tokens), kw_idx + 5)): # キーワードの後の4トークンをチェック
            tok = block_tokens[j] # 現在のトークン
            if "～" in tok or ":" in tok or "：" in tok: # 時間を示す可能性のある文字をチェック
                base_hours = tok; break # 基準営業時間を設定してループを抜ける

    name = "".join(name_tokens).replace(" ", "").replace("　", "") # 店舗名のトークンを結合して空白を削除
    return name, base_hours # 店舗名と基準営業時間を返す

def extract_status_tokens(block_tokens: List[str]) -> List[str]: # 営業可否トークンを抽出する関数
    statuses: List[str] = [] # 営業可否トークンを格納するリスト
    for t in block_tokens: # 各トークンを処理
        t_norm = t.replace(" ", "").replace("　", "") # トークンの空白を削除
        if TIME_PAIR_TOKEN_PATTERN.match(t_norm) or ONLY_MARK_TOKEN_PATTERN.match(t_norm): # 営業時間または営業可否のみのパターンをチェック
            statuses.append(t_norm) # マッチしたトークンをリストに追加
    if statuses and re.fullmatch(r"\d{1,2}", statuses[-1]): # 最後のトークンが日付のみの場合
        statuses = statuses[:-1]  # それを削除
    return statuses # 営業可否トークンリストを返す

def split_time_pair(token: str) -> Tuple[Optional[str], Optional[str]]: # 営業時間トークンを開始時間と終了時間に分割する関数
    m = TIME_PAIR_TOKEN_PATTERN.match(token) # トークンをパターンにマッチさせる
    if not m: return None, None # マッチしなかった場合はNoneを返す
    return m.group(1), m.group(2)   # 開始時間と終了時間を返す

def parse_pdf_to_rows(pdf_path: str) -> pd.DataFrame: # PDFを解析してデータフレームに変換する関数
    tokens = tokens_from_pdf(pdf_path) # PDFからトークンを抽出
    hits = find_building_hits(tokens) # 建物の境界を示すトークンを見つける
    blocks = split_blocks_by_hits(tokens, hits) # ヒットしたトークンでブロックを分割

    rows: List[Dict[str, Any]] = [] # データ行を格納するリスト
    for btoken, start, end, toks in blocks: # 各ブロックを処理
        place = btoken # ブロックの場所（建物名）
        shop_name, base_hours = extract_shop_name_and_base_hours(toks) # 店舗名と基準営業時間を抽出
        statuses = extract_status_tokens(toks) # 営業可否トークンを抽出
        for i, st in enumerate(statuses, start=1): # 各営業可否トークンを処理
            open_flag = "○" if st.startswith("○") else "×" # 営業可否を判定
            start_tm, end_tm = split_time_pair(st) if open_flag == "○" else (None, None) # 営業時間を分割
            rows.append( # データ行をリストに追加
                {
                    "日付": f"10月{i}日", # 日付（10月1日、10月2日、...）
                    "日": i, # 日（1、2、...）
                    "店舗名": shop_name, # 店舗名
                    "場所": place, # 場所（建物名）
                    "営業可否": open_flag, # 営業可否（○または×）
                    "営業開始": start_tm, # 営業開始時間
                    "営業終了": end_tm, # 営業終了時間
                    "基準営業時間": base_hours, # 基準営業時間
                }
            )

    return pd.DataFrame(rows) # データフレームを返す

# === ここまで：パーサ部 ===


# === ここから：レイアウト整形＆Excel書き出し（PDF風マトリクス） ===
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

def build_matrix(df: pd.DataFrame,
                 out_path: str = "10月_各店舗_営業マトリクス.xlsx",
                 year: int = 2025,
                 month: int = 10,
                 weekend_header_fill: str = "DDDDDD",
                 open_fill: str = "FFD966",
                 border_color: str = "000000",
                 highlight_days: Optional[list] = None):
    """
    df: parse_pdf_to_rows() の出力（縦長）
    out_path: 書き出しExcelパス
    year, month: ヘッダの曜日計算（2025/10 は水スタートでスクショに一致）
    weekend_header_fill: 土日ヘッダ色
    open_fill: 営業セルの塗り潰し色（薄い黄色）
    highlight_days: 特別日を列ハイライトしたいときに [8, 22, ...] のように指定（未指定なら無効）
    """
    if highlight_days is None:
        highlight_days = []

    # 1) 店舗軸の整備（No.付与）
    #   場所＋店舗名でユニーク化し、基準営業時間は最初の非NaNを採用
    shop_keys = (df["場所"].fillna("") + " / " + df["店舗名"].fillna("")).unique()
    # ただし順番はPDF順が理想だが、ここでは登場順（uniqueの順）
    key_to_no = {k: i+1 for i, k in enumerate(shop_keys)}

    # # 2) ピボットのための辞書（(key, day) -> cell text / flag）
    # def cell_text(row) -> str:
    #     if row["営業可否"] == "×":
    #         return "×"
    #     # ○ の場合
    #     if pd.notna(row["営業開始"]) and pd.notna(row["営業終了"]):
    #         # 改行で3段表示：○ / 開始 / 終了
    #         return f"○\n{row['営業開始']}\n{row['営業終了']}"
    #     return "○"

    def cell_text(row) -> str:
        """
        営業日の場合は「開始-終了」、休業日は「×」を表示
        """
        if row["営業可否"] == "×":
            return "×"
        if pd.notna(row["営業開始"]) and pd.notna(row["営業終了"]):
            return f"{row['営業開始']}-{row['営業終了']}"
        # 営業時間情報がない場合でも営業日なら基準営業時間を使う
        if pd.notna(row["基準営業時間"]):
            return row["基準営業時間"]
        return ""

    df = df.copy()
    df["key"] = df["場所"].fillna("") + " / " + df["店舗名"].fillna("")
    df["セル表示"] = df.apply(cell_text, axis=1)

    # 店舗ごとの属性（場所／名称／全営業時間リスト）
    agg = (df
        .sort_values(["key", "日"])
        .groupby("key")
        .agg({
            "場所": "first",
            "店舗名": "first",
            "基準営業時間": lambda x: ", ".join(sorted(set(v for v in x if pd.notna(v))))
        })
        .reset_index())


    # 3) ワークブック開始
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month}月"

    # 4) ヘッダ行構成
    #   Row1: タイトル
    #   Row2: 左側ラベル + 日付(1..31)
    #   Row3: 左側ラベル(空) + 曜日
    title = f"{month}月 各店舗 営業時間・日付別（○=営業）"
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4+31)
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers_left = ["No.", "場所", "名称", "営業時間"]
    for j, h in enumerate(headers_left, start=1):
        ws.cell(row=2, column=j, value=h)
        ws.cell(row=3, column=j, value="")  # 3行目はラベルなし

    # 日付と曜日
    _, last_day = calendar.monthrange(year, month)  # 10月は31日
    for d in range(1, last_day+1):
        col = 4 + d
        ws.cell(row=2, column=col, value=d)
        wd = date(year, month, d).weekday()  # Mon=0 ... Sun=6
        wd_jp = "月火水木金土日"[wd]
        ws.cell(row=3, column=col, value=wd_jp)

        # 土日ヘッダの色
        if wd >= 5:
            for r in (2,3):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=weekend_header_fill)

        # 特別日ハイライト（任意）
        if d in highlight_days:
            for r in (2,3):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="C6EFCE")  # 薄い緑

    # 5) データ行書き込み
    thin = Side(border_style="thin", color=border_color)
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    start_row = 4
    for i, row in agg.iterrows():
        r = start_row + i
        key = row["key"] if "key" in row else row["場所"]+" / "+row["店舗名"]
        no = key_to_no.get(key, i+1)

        # 左4列
        ws.cell(row=r, column=1, value=no)
        ws.cell(row=r, column=2, value=row["場所"])
        ws.cell(row=r, column=3, value=row["店舗名"])
        ws.cell(row=r, column=4, value=row["基準営業時間"])

        # 日付セル
        sub = df[df["key"] == key]
        day_map = {int(d): v for d, v in zip(sub["日"], sub["セル表示"])}

        for d in range(1, last_day+1):
            col = 4 + d
            val = day_map.get(d, "")
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            # 色付け：営業なら黄色、休業は白
            if isinstance(val, str) and len(val) > 0 and not val.startswith("×"):
                cell.fill = PatternFill("solid", fgColor=open_fill)

    # 6) 体裁：罫線・幅・固定
    max_row = start_row + len(agg) - 1
    max_col = 4 + last_day
    for rr in range(2, max_row+1):
        for cc in range(1, max_col+1):
            ws.cell(row=rr, column=cc).border = border_all
            ws.cell(row=rr, column=cc).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 列幅
    widths = {
        1: 4.0,   # No.
        2: 12.0,  # 場所
        3: 20.0,  # 名称
        4: 15.0,  # 営業時間
    }
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    for col_idx in range(5, max_col+1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 6.0

    # 行の高さ（見やすく）
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    for rr in range(4, max_row+1):
        ws.row_dimensions[rr].height = 22

    # ウィンドウ枠の固定（左4列＆ヘッダ2行を固定）
    ws.freeze_panes = "E4"

    # 7) 備考
    note_row = max_row + 2
    ws.cell(row=note_row, column=1, value="※ 食堂店舗のラストオーダーは閉店時間の30分前です")
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=max_col)

    # 保存
    wb.save(out_path)
    print(f"[OK] 10月＿各店舗＿営業.xlsxを出力しました: {Path(out_path).resolve()}")

    print("=== PDFパーサ開始 ===")
    pdf = Path(INPUT_PDF)
    if not pdf.exists():
        print(f"[ERROR] PDF が見つかりません: {pdf.resolve()}", file=sys.stderr)
        return
    
    # PDFを解析してDataFrameを取得
    df = parse_pdf_to_rows(str(pdf))
    
    # Excelファイルを生成（2025年10月のデータとして）
    build_matrix(
        df=df,
        out_path="10月_各店舗_営業マトリクス.xlsx",
        year=2025,
        month=10,
        highlight_days=[22]  # 現在の日付（10月22日）をハイライト
    )

def main():
    pdf = Path(INPUT_PDF)
    if not pdf.exists():
        print(f"[ERROR] PDF が見つかりません: {pdf.resolve()}", file=sys.stderr)
        return

    # PDFを解析してDataFrameを取得
    df = parse_pdf_to_rows(str(pdf))

    # Excelファイルを生成（2025年10月のデータとして）
    build_matrix(
        df=df,
        out_path="10月_各店舗_営業マトリクス.xlsx",
        year=2025,
        month=10,
        highlight_days=[22]  # 現在の日付（10月22日）をハイライト
    )

if __name__ == "__main__":
    main()
